from __future__ import annotations

from copy import copy

from openpyxl.styles import PatternFill, Font, Border, Alignment, Protection
from openpyxl.cell import Cell


class ExcelWriter:

    @staticmethod
    def copy_cell(source: Cell, target: Cell):

        target.value = source.value

        if source.has_style:

            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)
            target.number_format = source.number_format
            target.protection = copy(source.protection)

    @staticmethod
    def copy_row(sheet, source_row: int, target_row: int):

        for column in range(1, sheet.max_column + 1):

            ExcelWriter.copy_cell(

                sheet.cell(source_row, column),
                sheet.cell(target_row, column)

            )

        if source_row in sheet.row_dimensions:

            sheet.row_dimensions[target_row].height = (
                sheet.row_dimensions[source_row].height
            )

    @staticmethod
    def copy_column_widths(sheet):

        for key, value in sheet.column_dimensions.items():

            sheet.column_dimensions[key].width = value.width

    @staticmethod
    def copy_merged_cells(sheet):

        merged = list(sheet.merged_cells.ranges)

        for rng in merged:

            sheet.unmerge_cells(str(rng))
            sheet.merge_cells(str(rng))