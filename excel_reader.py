from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class ExcelTemplate:

    def __init__(self, template_path: str | Path):

        self.template_path = Path(template_path)

        if not self.template_path.exists():
            raise FileNotFoundError(self.template_path)

        self.workbook = None
        self.sheet: Worksheet | None = None

    def open(self):

        self.workbook = load_workbook(
            self.template_path
        )

        self.sheet = self.workbook.active

    @property
    def max_row(self):

        return self.sheet.max_row

    @property
    def max_column(self):

        return self.sheet.max_column

    def cell(self, row, column):

        return self.sheet.cell(
            row=row,
            column=column
        )

    def save(self, path):

        self.workbook.save(path)