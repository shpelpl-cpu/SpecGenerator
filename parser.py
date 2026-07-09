from __future__ import annotations

from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from models import InvoiceData, InvoiceItem
from utils import clean_cn, is_valid_cn, normalize_text, to_decimal


class InvoiceParser:

    REQUIRED_HEADERS = {
        "line": "line",
        "commodity code": "cn",
        "description": "description",
        "quantity": "quantity",
        "subtotal net weight(kgs)": "net_weight",
        "subtotal gross weight(kgs)": "gross_weight",
        "subtotal value": "value",
    }

    def __init__(self, file_path: str | Path):

        self.file_path = Path(file_path)

        if not self.file_path.exists():
            raise FileNotFoundError(self.file_path)

        self.workbook = None
        self.sheet = None

        self.header_row = None
        self.columns = {}

    def parse(self) -> InvoiceData:

        self._open()

        self._find_header()

        self._find_columns()

        return self._read_items()

    def _open(self):

        self.workbook = load_workbook(
            self.file_path,
            data_only=True,
        )

        self.sheet = self.workbook.active

    def _find_header(self):

        for row in self.sheet.iter_rows(max_row=50):

            values = [
                normalize_text(cell.value)
                for cell in row
            ]

            if (
                "line" in values
                and "commodity code" in values
                and "description" in values
            ):
                self.header_row = row[0].row
                return

        raise Exception("Nie znaleziono nagłówka tabeli.")

    def _find_columns(self):

        row = self.sheet[self.header_row]

        for cell in row:

            text = normalize_text(cell.value)

            if text in self.REQUIRED_HEADERS:

                self.columns[
                    self.REQUIRED_HEADERS[text]
                ] = cell.column

        missing = []

        for value in self.REQUIRED_HEADERS.values():

            if value not in self.columns:
                missing.append(value)

        if missing:

            raise Exception(
                f"Brak kolumn: {', '.join(missing)}"
            )

    def _read_items(self):

        data = InvoiceData()

        row = self.header_row + 1

        while True:

            line = self.sheet.cell(
                row=row,
                column=self.columns["line"]
            ).value

            if line is None:
                break

            cn = clean_cn(
                self.sheet.cell(
                    row=row,
                    column=self.columns["cn"]
                ).value
            )

            if not is_valid_cn(cn):
                break

            item = InvoiceItem(

                row_number=row,

                cn=cn,

                description_original=str(

                    self.sheet.cell(
                        row=row,
                        column=self.columns["description"]
                    ).value

                ).strip(),

                quantity=to_decimal(

                    self.sheet.cell(
                        row=row,
                        column=self.columns["quantity"]
                    ).value

                ),

                net_weight=to_decimal(

                    self.sheet.cell(
                        row=row,
                        column=self.columns["net_weight"]
                    ).value

                ),

                gross_weight=to_decimal(

                    self.sheet.cell(
                        row=row,
                        column=self.columns["gross_weight"]
                    ).value

                ),

                value=to_decimal(

                    self.sheet.cell(
                        row=row,
                        column=self.columns["value"]
                    ).value

                )

            )

            data.items.append(item)

            row += 1

        return data