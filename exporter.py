from __future__ import annotations

from copy import copy
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from models import GroupedItem


class Exporter:

    DATA_START_ROW = 2

    COL_LP = 1
    COL_CN = 2
    COL_QTY = 3
    COL_NET = 4
    COL_GROSS = 5
    COL_VALUE = 6
    COL_CURRENCY = 7
    COL_UNIT = 8
    COL_DESCRIPTION = 9
    COL_ORIGIN = 10
    COL_PREF = 11

    def __init__(self, template_path: str | Path):

        self.template_path = Path(template_path)

        if not self.template_path.exists():
            raise FileNotFoundError(self.template_path)

    @staticmethod
    def _number(value: Decimal) -> float:

        if value is None:
            return 0.0

        return float(value)

    @staticmethod
    def _quantity_text(value: Decimal) -> str:

        if value is None:
            return "0"

        return format(value.normalize(), "f")

    @staticmethod
    def _copy_style(source, target):

        target._style = copy(source._style)

    @staticmethod
    def _move_row_dimensions(
        sheet: Worksheet,
        start_row: int,
        amount: int,
    ):

        """Move row metadata which openpyxl does not move with cells."""

        rows_to_move = [
            row
            for row in sheet.row_dimensions
            if row >= start_row
        ]

        for row in sorted(rows_to_move, reverse=True):

            dimension = copy(sheet.row_dimensions[row])
            dimension.index = row + amount

            sheet.row_dimensions[row + amount] = dimension
            del sheet.row_dimensions[row]

    @staticmethod
    def _move_merged_cells(
        sheet: Worksheet,
        start_row: int,
        amount: int,
    ):

        """Keep merged ranges aligned with the rows moved by insert_rows."""

        merged_ranges = list(sheet.merged_cells.ranges)

        sheet.merged_cells.ranges.clear()

        for merged_range in merged_ranges:

            if merged_range.min_row >= start_row:
                merged_range.shift(row_shift=amount)

            elif merged_range.max_row >= start_row:
                merged_range.max_row += amount

            sheet.merged_cells.add(merged_range)
            merged_range.format()

    @staticmethod
    def _copy_data_row_merged_cells(
        sheet: Worksheet,
        template_row: int,
        target_rows: range,
    ):

        """Copy horizontal merged ranges belonging to the template row."""

        template_ranges = [
            merged_range
            for merged_range in sheet.merged_cells.ranges
            if (
                merged_range.min_row == template_row
                and merged_range.max_row == template_row
            )
        ]

        for row in target_rows:

            for merged_range in template_ranges:
                sheet.merge_cells(
                    start_row=row,
                    start_column=merged_range.min_col,
                    end_row=row,
                    end_column=merged_range.max_col,
                )

    @staticmethod
    def _capture_print_settings(sheet: Worksheet) -> dict:

        """Keep print settings intact while rows are inserted."""

        return {
            "page_setup": copy(sheet.page_setup),
            "page_margins": copy(sheet.page_margins),
            "print_options": copy(sheet.print_options),
            "print_area": sheet.print_area,
            "print_title_rows": sheet.print_title_rows,
            "print_title_cols": sheet.print_title_cols,
        }

    @staticmethod
    def _restore_print_settings(
        sheet: Worksheet,
        settings: dict,
    ):

        sheet.page_setup = copy(settings["page_setup"])
        sheet.page_margins = copy(settings["page_margins"])
        sheet.print_options = copy(settings["print_options"])

        if settings["print_area"] is not None:
            sheet.print_area = settings["print_area"]

        if settings["print_title_rows"] is not None:
            sheet.print_title_rows = settings["print_title_rows"]

        if settings["print_title_cols"] is not None:
            sheet.print_title_cols = settings["print_title_cols"]

    def _insert_rows(
        self,
        sheet: Worksheet,
        amount: int,
    ):

        if amount <= 0:
            return

        insert_at = self.DATA_START_ROW + 1
        template_row = self.DATA_START_ROW
        max_column = sheet.max_column

        self._move_row_dimensions(
            sheet,
            insert_at,
            amount,
        )

        sheet.insert_rows(
            insert_at,
            amount
        )

        self._move_merged_cells(
            sheet,
            insert_at,
            amount,
        )

        inserted_rows = range(
            insert_at,
            insert_at + amount,
        )

        for row in inserted_rows:

            template_dimension = copy(
                sheet.row_dimensions[template_row]
            )
            template_dimension.index = row
            sheet.row_dimensions[row] = template_dimension

            for col in range(
                1,
                max_column + 1
            ):

                src = sheet.cell(template_row, col)
                dst = sheet.cell(row, col)

                self._copy_style(src, dst)

        self._copy_data_row_merged_cells(
            sheet,
            template_row,
            inserted_rows,
        )

    def _write_row(
        self,
        sheet: Worksheet,
        row: int,
        lp: int,
        group: GroupedItem,
    ):

        sheet.cell(row, self.COL_LP).value = lp

        sheet.cell(row, self.COL_CN).value = group.cn

        sheet.cell(row, self.COL_QTY).value = 0

        sheet.cell(row, self.COL_NET).value = self._number(
            group.net_weight
        )

        sheet.cell(row, self.COL_GROSS).value = self._number(
            group.gross_weight
        )

        sheet.cell(row, self.COL_VALUE).value = self._number(
            group.value
        )

        sheet.cell(row, self.COL_CURRENCY).value = "EUR"

        sheet.cell(row, self.COL_UNIT).value = self._number(
            group.quantity
        )

        sheet.cell(
            row,
            self.COL_DESCRIPTION
        ).value = (
            f"{group.description} "
            f"({self._quantity_text(group.quantity)} szt)"
        )

        sheet.cell(
            row,
            self.COL_ORIGIN
        ).value = "CN"

        sheet.cell(
            row,
            self.COL_PREF
        ).value = "100"


    def export(
        self,
        groups: list[GroupedItem],
        output_path: str | Path,
    ):

        workbook = load_workbook(
            self.template_path
        )

        sheet = workbook.active
        print_settings = self._capture_print_settings(sheet)

        if len(groups) > 1:

            self._insert_rows(
                sheet,
                len(groups) - 1
            )

        row = self.DATA_START_ROW

        lp = 1

        for group in groups:

            self._write_row(
                sheet,
                row,
                lp,
                group
            )

            row += 1
            lp += 1

        output_path = Path(output_path)

        output_path.parent.mkdir(
            parents=True,
            exist_ok=True
        )

        self._restore_print_settings(
            sheet,
            print_settings,
        )

        workbook.save(output_path)

        workbook.close()
