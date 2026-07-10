from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import load_workbook

from exporter import Exporter
from grouper import InvoiceGrouper
from mapper import DescriptionMapper
from parser import InvoiceParser
from validator import InvoiceValidator


ROOT_DIR = Path(__file__).resolve().parents[1]
APPROVED_FILE = Path(
    r"C:\Users\towar\OneDrive\Pulpit\test\Specyfikacja 87644030560 (1).xlsx"
)


def specification_rows(path: Path) -> list[tuple]:

    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    rows = []

    for row in range(2, sheet.max_row + 1):
        cn = sheet.cell(row, 2).value
        if cn is None:
            continue

        rows.append(
            (
                str(cn),
                Decimal(str(sheet.cell(row, 8).value)),
                Decimal(str(sheet.cell(row, 4).value)),
                Decimal(str(sheet.cell(row, 5).value)),
                Decimal(str(sheet.cell(row, 6).value)),
                " ".join(str(sheet.cell(row, 9).value).split()),
            )
        )

    workbook.close()
    return rows


class ApprovedSpecificationTest(unittest.TestCase):

    def test_generated_specification_matches_approved_file(self):

        self.assertTrue(APPROVED_FILE.is_file())
        invoice = InvoiceParser(ROOT_DIR / "invoice.xlsx").parse()
        mapper = DescriptionMapper(ROOT_DIR / "mappings.json")

        for item in invoice.items:
            item.description_pl = mapper.map(
                item.description_original,
                item.gender,
                item.cn,
            )

        groups = InvoiceGrouper().group(invoice)
        validation = InvoiceValidator().validate(invoice, groups)

        self.assertTrue(validation.success, "\n".join(validation.errors))
        self.assertEqual(353, len(invoice.items))
        self.assertEqual(44, len(invoice.unique_cn))
        self.assertEqual(49, len(groups))
        self.assertEqual(Decimal("8471"), invoice.total_quantity)
        self.assertEqual(Decimal("1526.660"), invoice.total_net_weight)
        self.assertEqual(Decimal("1606.900"), invoice.total_gross_weight)
        self.assertEqual(Decimal("23594.38"), invoice.total_value)

        with TemporaryDirectory() as directory:
            output_path = Path(directory) / "specification.xlsx"
            Exporter(ROOT_DIR / "templates" / "spec_template.xlsx").export(
                groups,
                output_path,
            )

            self.assertEqual(
                specification_rows(APPROVED_FILE),
                specification_rows(output_path),
            )

            workbook = load_workbook(output_path, data_only=True)
            self.assertEqual(50, workbook.active.max_row)
            workbook.close()
